import openpyxl
from utils import get_col_index, compare_data, print_pretty_differences, export_differences_to_txt
import shutil
import os
import json
from config import LIST_KEY_FILE1, LIST_KEY_FILE2, MAPPING_FIELD, BASE_FOLDER, BASE_FOLDER_1, BASE_FOLDER_2

def read_excel_cells(file_path, list_keys, min_row=7):
    """
    Reads all cells from the specified sheet in an Excel (.xlsx) file.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read.
    :return: A list of lists containing the cell values.
    """
    data = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        for sheet in sheet_names:
            if any(keyword in sheet.upper() for keyword in ["- HK", "WELL", "TATC"]):
                continue
            sheet_data = workbook[sheet]
            sheet_result = []

            for row in sheet_data.iter_rows(values_only=True, min_row=min_row):
                user = {item["key"]: row[get_col_index(item["value"])] for item in list_keys}
                sheet_result.append(user)
            data[sheet] = sheet_result

        return data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None
    
def compare_excel_files(file1_data, file2_data, mapping, file_name):
    list_sheets1 = read_excel_cells(file1_data["file_path"], file1_data["list_keys"], file1_data["min_row"])
    list_sheets2 = read_excel_cells(file2_data["file_path"], file2_data["list_keys"], file2_data["min_row"])
    # Export the data to JSON files
    os.makedirs(f"json/{BASE_FOLDER_1}", exist_ok=True)
    os.makedirs(f"json/{BASE_FOLDER_2}", exist_ok=True)
    with open(f"json/{BASE_FOLDER_1}/{file_name}.json", "w", encoding="utf-8") as f1:
        json.dump(list_sheets1, f1, ensure_ascii=False, indent=4)

    with open(f"json/{BASE_FOLDER_2}/{file_name}.json", "w", encoding="utf-8") as f2:
        json.dump(list_sheets2, f2, ensure_ascii=False, indent=4)
    
    # Compare the two data objects
    differences = compare_data(list_sheets1, list_sheets2, mapping)
    
    print(f"===================> So sánh Lớp {file_name} <===================")
    if not differences:
        print(f"===================> Lớp {file_name} ĐÚNG <===================")
    else:
        print_pretty_differences(differences)
    
    # Export the differences to a .txt file
    output_file = f"outputs/{file_name}.txt"
    export_differences_to_txt(differences, output_file)
    
    
if __name__ == "__main__":

          
    folder1 = os.path.join(BASE_FOLDER, BASE_FOLDER_1)
    folder2 = os.path.join(BASE_FOLDER, BASE_FOLDER_2)
    if os.path.exists("outputs"):
        shutil.rmtree("outputs")
    os.makedirs("outputs", exist_ok=True)
    if os.path.isdir(folder1) and os.path.isdir(folder2):
        files1 = [f for f in os.listdir(folder1) if f.endswith(".xlsx") and not f.startswith("~$")]
        files2 = [f for f in os.listdir(folder2) if f.endswith(".xlsx") and not f.startswith("~$")]

        for file_name in files1:
            if file_name in files2:
                file1 = os.path.join(folder1, file_name)
                file2 = os.path.join(folder2, file_name)
                
                file_name_without_ext = file_name.split(".")[0]
                
                file1_data = {
                    "list_keys": LIST_KEY_FILE1,
                    "file_path": file1,
                    "min_row": 9,
                }
                file2_data = {
                    "list_keys": LIST_KEY_FILE2,
                    "file_path": file2,
                    "min_row": 7,
                }
                compare_excel_files(file1_data, file2_data, MAPPING_FIELD, file_name_without_ext)
